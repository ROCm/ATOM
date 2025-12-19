import os
import sys
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_log_file(log_file_path):
    """解析日志文件，提取关键指标"""
    results = []
    
    with open(log_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    test_blocks = re.findall(
        r'开始测试 ISL=(\d+), OSL=(\d+), CONC=(\d+)(.*?)==================================================',
        content,
        re.DOTALL
    )
    
    for isl, osl, conc, block in test_blocks:
        mean_ttft = re.search(r'Mean TTFT \(ms\):\s+([\d.]+)', block)
        mean_tpot = re.search(r'Mean TPOT \(ms\):\s+([\d.]+)', block)
        total_throughput = re.search(r'Total Token throughput \(tok/s\):\s+([\d.]+)', block)
        
        if mean_ttft and mean_tpot and total_throughput:
            results.append({
                'ISL': int(isl),
                'OSL': int(osl),
                'CONC': int(conc),
                'Mean TTFT (ms)': float(mean_ttft.group(1)),
                'Mean TPOT (ms)': float(mean_tpot.group(1)),
                'Total Throughput (tok/s)': float(total_throughput.group(1))
            })
    
    return results

def create_excel_repeating_format(results, output_file):
    """创建重复格式的Excel文件：每个ISL/OSL组合重复显示，确保ISL行有完整边框"""
    
    # 按ISL和CONC排序
    results.sort(key=lambda x: (x['ISL'], -x['CONC']))
    
    # 创建新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Results"
    
    # 定义边框样式 - 确保所有边框都设置
    full_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义填充样式
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    isl_header_fills = [
        PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
    ]
    
    # 定义字体样式
    header_font = Font(bold=True, size=11, color="000000")
    isl_title_font = Font(bold=True, size=12, color="000000")
    data_font = Font(size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 1
    grouped_results = {}
    for result in results:
        key = (result['ISL'], result['OSL'])
        if key not in grouped_results:
            grouped_results[key] = []
        grouped_results[key].append(result)
    
    # 按ISL排序
    sorted_keys = sorted(grouped_results.keys())
    
    # 为每个ISL/OSL组合创建独立的4行区域
    for idx, (isl, osl) in enumerate(sorted_keys):
        data_list = grouped_results[(isl, osl)]
        data_list.sort(key=lambda x: -x['CONC'])  # 按CONC降序排列
        
        # 第1行：ISL/OSL标题行（合并4个单元格） - 确保有完整边框
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        
        cell = ws.cell(row=current_row, column=1, value=f"ISL={isl}, OSL={osl}")
        cell.fill = isl_header_fills[idx % len(isl_header_fills)]
        cell.font = isl_title_font
        cell.alignment = center_alignment
        cell.border = full_border  # 确保有完整边框
        
        # 为合并单元格的每个单元格单独设置边框（确保显示）
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).border = full_border
        
        current_row += 1
        
        # 第2行：表头行
        for col, header in enumerate(["Concurrency", "Mean TTFT (ms)", "Mean TPOT (ms)", "Total Throughput (tok/s)"], 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = full_border
        
        current_row += 1
        
        # 第3-4行：数据行（最多2个并发数）
        for i, data in enumerate(data_list):  # 只取前2个并发数
            # Concurrency
            cell = ws.cell(row=current_row, column=1, value=data['CONC'])
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = full_border
            
            # Mean TTFT
            cell = ws.cell(row=current_row, column=2, value=data['Mean TTFT (ms)'])
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = full_border
            
            # Mean TPOT
            cell = ws.cell(row=current_row, column=3, value=data['Mean TPOT (ms)'])
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = full_border
            
            # Total Throughput
            cell = ws.cell(row=current_row, column=4, value=data['Total Throughput (tok/s)'])
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = full_border
            
            current_row += 1
        
        # 添加空行分隔不同ISL/OSL组合（不带边框）
        current_row += 1
    
    # 删除最后的空行
    if current_row > 1:
        current_row -= 1
    
    # 自动调整列宽
    for col_idx in range(1, 5):
        column_letter = get_column_letter(col_idx)
        
        if col_idx == 1:  # Concurrency列
            ws.column_dimensions[column_letter].width = 12
        elif col_idx == 2 or col_idx == 3:  # Mean TTFT/TPOT列
            ws.column_dimensions[column_letter].width = 15
        else:  # Total Throughput列
            ws.column_dimensions[column_letter].width = 22
    
    # 保存文件
    wb.save(output_file)
    return output_file

def print_csv_repeating_format(results):
    """以CSV表格格式打印数据（重复格式）"""
    
    # 按ISL分组
    grouped_results = {}
    for result in results:
        key = (result['ISL'], result['OSL'])
        if key not in grouped_results:
            grouped_results[key] = []
        grouped_results[key].append(result)
    
    # 按ISL排序
    sorted_keys = sorted(grouped_results.keys())
    
    print("\n" + "=" * 60)
    print("DEEPSEEK-V3.2 BENCHMARK RESULTS")
    print("=" * 60)
    
    # 为每个ISL/OSL组合打印独立的表格
    for idx, (isl, osl) in enumerate(sorted_keys):
        data_list = grouped_results[(isl, osl)]
        data_list.sort(key=lambda x: -x['CONC'])  # 按CONC降序排列
        
        print(f"\n┌{'ISL=' + str(isl) + ', OSL=' + str(osl):^52}┐")
        print("├────────────────────────────────────────────────────┤")
        print("│ Concurrency │ Mean TTFT │ Mean TPOT │  Throughput  │")
        print("├─────────────┼───────────┼───────────┼──────────────┤")
        
        # 打印最多2行数据
        for i, data in enumerate(data_list):
            print(f"│ {data['CONC']:^11} │ {data['Mean TTFT (ms)']:>9.2f} │ {data['Mean TPOT (ms)']:>9.2f} │ {data['Total Throughput (tok/s)']:>12.2f} │")
        
        # 如果数据少于2行，用空行补齐
        while len(data_list) < 2 and i < 1:
            print("│             │           │           │              │")
            i += 1
        
        print("└─────────────┴───────────┴───────────┴──────────────┘")
    
    print("=" * 60)

def main():
    # 输入文件路径
    parser = argparse.ArgumentParser(
        description='benchmark infomation',
    )
    
    # 添加参数
    parser.add_argument(
        'input_file',
        nargs='?',  # 可选参数
        default='result.txt',
        help='the path of input file, default:result.txt.'
    )
    
    args = parser.parse_args()
    log_file_path = args.input_file
    
    if not os.path.isfile(log_file_path):
        print(f"file not exist - {log_file_path}")
        sys.exit(1)

    output_file = "benchmark.xlsx"
    
    try:
        # 解析日志文件
        results = parse_log_file(log_file_path)
        
        if results:
            # 以CSV表格格式打印数据
            print_csv_repeating_format(results)
            
            # 创建Excel文件
            create_excel_repeating_format(results, output_file)
            
        else:
            print("没有找到有效的基准测试结果")
            
    except FileNotFoundError:
        print(f"错误: 找不到日志文件 {log_file_path}")
    except Exception as e:
        print(f"处理过程中出现错误: {e}")

if __name__ == "__main__":
    main()