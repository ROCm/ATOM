#!/usr/bin/env python3
"""Build LLM-Booster dashboard data for ATOM kernel breakdown profiles."""

from __future__ import annotations

import argparse
import json
import re
import shutil
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-") or "unknown"


def normalize_precision(*values: str | None) -> str | None:
    merged = " ".join(str(v or "") for v in values).lower()
    if "bf16" in merged:
        return "bf16"
    if "mxfp4" in merged or "fp4" in merged:
        return "fp4"
    if "fp8" in merged:
        return "fp8"
    return None


def client_label(raw: str | None) -> str:
    value = (raw or "").strip().lower()
    return "vllm bench" if value in {"vllm bench", "vllmbench"} else "inferencemax"


def build_config_label(manifest: dict[str, Any]) -> str:
    model = manifest.get("dashboard_model") or manifest.get("model") or "unknown"
    label = f"atom-vllm_{slugify(str(model))}"
    precision = normalize_precision(manifest.get("precision"), manifest.get("dashboard_model"), manifest.get("model"))
    if precision:
        label += f"_{precision}"
    tp = manifest.get("tensor_parallel_size")
    try:
        tp_value = int(tp) if tp is not None else None
    except (TypeError, ValueError):
        tp_value = None
    if tp_value:
        label += f"_tp{tp_value}"
    if client_label(manifest.get("benchmark_client")) == "inferencemax":
        label += "_inferencemax"
    return label


def _as_float(value: Any) -> float | None:
    try:
        if value in ("", None):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _as_int(value: Any) -> int | None:
    try:
        if value in ("", None):
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_detail_xlsx(xlsx_path: Path) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)

    kernel_summary: list[dict[str, Any]] = []
    grouped_sheets: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == "kernel_summary":
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(v not in (None, "") for v in row):
                    continue
                kernel_summary.append(
                    {
                        "gpu_kernel": row[0],
                        "calls": _as_int(row[1]),
                        "total_duration_us": _as_float(row[2]),
                        "avg_duration_us": _as_float(row[3]),
                        "pct": _as_float(row[4]),
                    }
                )
            continue

        groups: list[dict[str, Any]] = []
        current_group: dict[str, Any] | None = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            values = list(row[:5]) + [None] * max(0, 5 - len(row))
            if not any(v not in (None, "") for v in values):
                continue

            first = values[0]
            rest = values[1:5]

            if first == "TOTAL":
                if current_group is not None:
                    current_group["total_duration_us"] = _as_float(values[2])
                    current_group["total_pct"] = _as_float(values[3])
                continue

            if first not in (None, "") and not any(v not in (None, "") for v in rest):
                current_group = {
                    "label": str(first),
                    "rows": [],
                    "total_duration_us": None,
                    "total_pct": None,
                }
                groups.append(current_group)
                continue

            if current_group is None:
                current_group = {
                    "label": "",
                    "rows": [],
                    "total_duration_us": None,
                    "total_pct": None,
                }
                groups.append(current_group)

            current_group["rows"].append(
                {
                    "cpu_module": first or "",
                    "gpu_kernel": values[1] or "",
                    "duration_us": _as_float(values[2]),
                    "pct": _as_float(values[3]),
                    "input_shapes": values[4] or "",
                }
            )

        grouped_sheets[sheet_name] = groups

    return kernel_summary, grouped_sheets


def build_detail_payload(
    manifest: dict[str, Any],
    xlsx_relative_path: Path,
    kernel_summary: list[dict[str, Any]],
    grouped_sheets: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    model = manifest.get("dashboard_model") or manifest.get("model") or "unknown"
    precision = normalize_precision(
        manifest.get("precision"),
        manifest.get("dashboard_model"),
        manifest.get("model"),
    )
    tp = _as_int(manifest.get("tensor_parallel_size"))
    isl = _as_int(manifest.get("isl"))
    osl = _as_int(manifest.get("osl"))
    ratio = _as_float(manifest.get("random_range_ratio"))
    client_bench = client_label(manifest.get("benchmark_client"))

    return {
        "profile_id": manifest.get("profile_id"),
        "date": manifest.get("date"),
        "timestamp": _as_int(manifest.get("timestamp")),
        "model": model,
        "backend": manifest.get("backend") or "vLLM-ATOM",
        "config_label": build_config_label(manifest),
        "precision": precision,
        "source": manifest.get("source") or "ATOM Dashboard",
        "client_bench": client_bench,
        "hardware": manifest.get("hardware") or "mi355x",
        "isl": isl,
        "osl": osl,
        "isl_osl": f"{isl}|{osl}" if isl is not None and osl is not None else None,
        "concurrency": _as_int(manifest.get("concurrency")),
        "ratio": ratio,
        "num_prompts": _as_int(manifest.get("num_prompts")),
        "tp": str(tp) if tp is not None else None,
        "tensor_parallel_size": tp,
        "profile_layer": _as_int(manifest.get("profile_layer")),
        "profile_percentile": _as_float(manifest.get("profile_percentile")),
        "run_url": manifest.get("run_url"),
        "atom_source_repository": manifest.get("atom_source_repository"),
        "atom_source_ref": manifest.get("atom_source_ref"),
        "atom_source_sha": manifest.get("atom_source_sha"),
        "gpu_name": manifest.get("gpu_name"),
        "oot_image_tag": manifest.get("oot_image_tag"),
        "rocm_version": manifest.get("rocm_version"),
        "model_id": manifest.get("model_id"),
        "xlsx_file": xlsx_relative_path.as_posix(),
        "kernel_summary": kernel_summary,
        "sheets": grouped_sheets,
    }


def rebuild_index(data_dir: Path) -> dict[str, Any]:
    profiles: list[dict[str, Any]] = []
    for detail_path in sorted(data_dir.glob("*.json")):
        if detail_path.name == "index.json":
            continue
        try:
            detail = json.loads(detail_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeDecodeError, json.JSONDecodeError):
            continue

        kernel_summary = detail.get("kernel_summary") or []
        sheet_names = list((detail.get("sheets") or {}).keys())
        profiles.append(
            {
                "profile_id": detail.get("profile_id", detail_path.stem),
                "date": detail.get("date"),
                "timestamp": detail.get("timestamp"),
                "model": detail.get("model"),
                "backend": detail.get("backend"),
                "config_label": detail.get("config_label"),
                "precision": detail.get("precision"),
                "source": detail.get("source"),
                "client_bench": detail.get("client_bench"),
                "hardware": detail.get("hardware"),
                "isl": detail.get("isl"),
                "osl": detail.get("osl"),
                "concurrency": detail.get("concurrency"),
                "ratio": detail.get("ratio"),
                "num_prompts": detail.get("num_prompts"),
                "tp": detail.get("tp"),
                "tensor_parallel_size": detail.get("tensor_parallel_size"),
                "profile_layer": detail.get("profile_layer"),
                "profile_percentile": detail.get("profile_percentile"),
                "run_url": detail.get("run_url"),
                "xlsx_file": detail.get("xlsx_file"),
                "n_kernel_summary_rows": len(kernel_summary),
                "sheet_names": sheet_names,
                "file": detail_path.name,
            }
        )

    profiles.sort(
        key=lambda item: (
            item.get("timestamp") or 0,
            item.get("date") or "",
            item.get("profile_id") or "",
        ),
        reverse=True,
    )
    index = {
        "lastUpdate": int(time.time() * 1000),
        "source": "ATOM kernel breakdown profiles",
        "profiles": profiles,
    }
    (data_dir / "index.json").write_text(json.dumps(index, indent=2) + "\n", encoding="utf-8")
    return index


def build_dashboard_data(artifact_dir: Path, dashboard_dir: Path) -> None:
    data_dir = dashboard_dir / "data" / "kernel-breakdown"
    files_dir = data_dir / "files"
    data_dir.mkdir(parents=True, exist_ok=True)
    files_dir.mkdir(parents=True, exist_ok=True)

    manifest_paths = (
        sorted(artifact_dir.rglob("*.kernel_breakdown_manifest.json"))
        if artifact_dir.exists()
        else []
    )
    written = 0

    for manifest_path in manifest_paths:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        xlsx_name = manifest.get("xlsx_file")
        if not xlsx_name:
            print(f"Skipping {manifest_path.name}: missing xlsx_file")
            continue

        xlsx_path = manifest_path.parent / xlsx_name
        if not xlsx_path.is_file():
            print(f"Skipping {manifest_path.name}: missing {xlsx_name}")
            continue

        profile_id = str(manifest.get("profile_id") or manifest_path.stem)
        target_xlsx_path = files_dir / f"{profile_id}.xlsx"
        shutil.copy2(xlsx_path, target_xlsx_path)

        kernel_summary, grouped_sheets = parse_detail_xlsx(target_xlsx_path)
        detail = build_detail_payload(
            manifest,
            target_xlsx_path.relative_to(dashboard_dir / "data"),
            kernel_summary,
            grouped_sheets,
        )
        detail_path = data_dir / f"{profile_id}.json"
        detail_path.write_text(json.dumps(detail, indent=2) + "\n", encoding="utf-8")
        written += 1

    index = rebuild_index(data_dir)
    print(
        f"Wrote/updated {written} kernel breakdown profiles; "
        f"index now lists {len(index['profiles'])} profile(s)."
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--artifact-dir", required=True, type=Path)
    parser.add_argument("--dashboard-dir", required=True, type=Path)
    args = parser.parse_args()

    build_dashboard_data(args.artifact_dir, args.dashboard_dir)


if __name__ == "__main__":
    main()
